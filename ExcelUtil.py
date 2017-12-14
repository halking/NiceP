#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2017/11/22 18:35
# @Author  : Steven
# @Site    : 
# @File    : ExcelUtil.py
# @Software: PyCharm Community Edition

from openpyxl import Workbook,load_workbook
import os


def readFile(rootDir):
    dirs = os.listdir(rootDir)
    for file in dirs:
        if file.endswith('.xlsx'):
            print(rootDir+file)
            readExcel(rootDir + file)

def readExcel(fileName,defaultSheet='All-sheet'):
    wb = load_workbook(fileName)
    writeBook = Workbook()
    writeSheet = writeBook.create_sheet(defaultSheet)
    max_row = 0;
    for sheet in wb:
        rows = sheet.iter_rows()
        for row in rows:
            max_row+=1
            max_col = 0
            for cell in row:
                max_col+=1
                writeSheet.cell(row=max_row,column=max_col,value=cell.value)
    writeExcel(writeBook)

def writeExcel(workBook,fileName='default_excel.xlsx'):
    workBook.save(fileName)

# readExcel("E:/chatbot/test.xlsx")
path = 'E:/chatbot/'
readFile(path);
# defaultSheet='All-sheet'
# fileName='default_excel.xlsx'
# writeBook = Workbook()
# writeSheet = writeBook.create_sheet(defaultSheet)
# for row in range(10, 20):
#      for col in range(27, 54):
#         writeSheet.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
# writeExcel(writeBook,path+fileName)
#
# for row in range(10, 20):
#      for col in range(27, 54):
#          print(row,col)
#          print(get_column_letter(col))